import pandas as pd
import json
import time
from sklearn.preprocessing import MinMaxScaler
import ssl
import certifi
from google import genai
import streamlit as st
from io import BytesIO
from pathlib import Path
import requests
import os
import concurrent.futures
from langchain.agents import Tool
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import base64
from datetime import datetime

API_KEYS = [
    "AIzaSyD09_gws5tBYZmD0YHF1etSZ7K-7wePIh0",
    "AIzaSyBJ2N1RHTTTQMXUod7jPymZwbgnPsdgLsY",
    "AIzaSyBwvI4kSZWOnWG3Km6kpUbqD87wIUVcoHs",
    "AIzaSyDKvI5lwfrihbcXXaXxaQWhGULE77afyrg",
    "AIzaSyCxpTPYFq91HfeUVqe8JD3RjiU4nV63WH8",
    "AIzaSyCWZVz-ciOp91vKr2u7J87IktK2skygOro",
    "AIzaSyB11u1-TTuvIRNhSAp44PgWWpoK9kq1mAo",
    "AIzaSyBxusefsMEbKv6HAoYxECpOIqbKO-pCs2g",
    "AIzaSyDIAYd4QdTBQO4MVOnAvoA5tNEozVYdflE",
    "AIzaSyBw6zUcIsp5t4QZxI_BRiPphYJzf7mq8p4",
    "AIzaSyC3EpZaqKLQwxCGUxKLzuwzvtKT2EjYTEA",
    "AIzaSyAkXdS9nAA35pdOX4kZQaFOgOznjU9MlDs",
    "AIzaSyBZqnpTMHL8Zap2CIrqifqXVA5YB30Apuw",
    "AIzaSyBqTtltNANsAhbodnxfFJOFq8vaGszJPqQ",
    "AIzaSyCC2RTsg8ArBgXj8t82-w-agFE82s0CUHw",
    "AIzaSyDvtLtNuVVlgNBvzwPRl42RyWZJqRsCI4Q",
    "AIzaSyATYlQN6L7SJz7mY7wScnyB8G_DqRsJQT4",
    "AIzaSyBW8Q1amjzs0_XLHaKaecyZuQJe0U5qhZU",
    "AIzaSyA7YtWUSsljlQuWOuy3fSBajot2rI5D3e8",
    "AIzaSyAsFagF5Z-A_o2pvUiAwpzqXpDpRNjhwfM",
    "AIzaSyDG8LTKH4NGqQcaGAz76z4hKAQ95jVjz4c",
    "AIzaSyDwB9W3SJjG5qkTd58L8ToX0xmi57Kh8d4",
    "AIzaSyBNAb6TSR4mhq82WtW2wHSCOUDK73IDbfs",
    "AIzaSyB51i5YnENFBE8aYncinPtwLk1dThl2CuA"
]

def load_font(font_path):
    """Load font file and convert to base64"""
    if os.path.exists(font_path):
        with open(font_path, "rb") as f:
            font_data = f.read()
        return base64.b64encode(font_data).decode()
    return None

# Load Nazanin fonts from your repository
font_regular = load_font("0 Nazanin.TTF")
font_bold = load_font("0 Nazanin Bold.TTF")

if font_regular and font_bold:
    font_css = f"""
    <style>
      @font-face {{
        font-family: 'Nazanin';
        src: url(data:font/truetype;charset=utf-8;base64,{font_regular}) format('truetype');
        font-weight: normal;
        font-style: normal;
      }}

      @font-face {{
        font-family: 'Nazanin';
        src: url(data:font/truetype;charset=utf-8;base64,{font_bold}) format('truetype');
        font-weight: bold;
        font-style: normal;
      }}

      html, body, [class^="st-"], [class*=" st-"], .block-container {{
        font-family: 'Nazanin', Tahoma, sans-serif !important;
        direction: rtl !important;
        text-align: right !important;
      }}
      
      /* Target all Streamlit elements */
      .stButton > button {{
        font-family: 'Nazanin', Tahoma, sans-serif !important;
      }}
      
      .stSelectbox label, .stMultiSelect label, .stTextInput label {{
        font-family: 'Nazanin', Tahoma, sans-serif !important;
      }}
      
      div[data-testid="stDataFrame"] * {{
        font-family: 'Nazanin', Tahoma, sans-serif !important;
      }}
    </style>
    """
    st.markdown(font_css, unsafe_allow_html=True)
else:
    st.warning("فونت Nazanin یافت نشد. از فونت پیش‌فرض استفاده می‌شود.")


def style_excel(path): 
    wb = openpyxl.load_workbook(path) 
    ws = wb.active 

    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_even = PatternFill(start_color="EAF3FA", end_color="EAF3FA", fill_type="solid")

    header_font = Font(bold=True, name='B Homa', size=14)
    row_font = Font(name='B Homa', size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(border_style="thin", color="CCCCCC"),
        right=Side(border_style="thin", color="CCCCCC"),
        top=Side(border_style="thin", color="CCCCCC"),
        bottom=Side(border_style="thin", color="CCCCCC"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        fill = row_fill_even if idx % 2 == 0 else row_fill_odd
        for cell in row:
            cell.fill = fill
            cell.font = row_font
            cell.alignment = center_align
            cell.border = border

    for col in ws.columns:
        if col[0].value == "تحلیل نهایی":
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for col in ws.columns: 
        max_length = 0 
        column = col[0].column_letter 
        for cell in col: 
            try: 
                if cell.value:
                    max_length = max(max_length, len(str(cell.value))) 
            except: 
                pass 
        adjusted_width = min(max_length + 3, 50) 
        ws.column_dimensions[column].width = adjusted_width 

    ws.freeze_panes = ws["A2"] 

    wb.save(path)


class RotatingGeminiLLM:
    def __init__(self, api_keys, model="gemini-2.5-flash"):
        self.api_keys = api_keys
        self.model = model
        self.idx = 0

    def invoke(self, messages):
        num_keys = len(self.api_keys)
        start_idx = self.idx
        for i in range(num_keys):
            api_key = self.api_keys[self.idx]
            llm = ChatGoogleGenerativeAI(model=self.model, google_api_key=api_key)
            try:
                result = llm.invoke(messages)
                return result
            except Exception as e:
                print(f"⚠️ خطا با API {api_key[:10]}...: {str(e)}")
                self.idx = (self.idx + 1) % num_keys
                if self.idx == start_idx:
                    raise RuntimeError("❌ تمام API Keyها با خطا مواجه شدند.")
        raise RuntimeError("❌ تمام API Keyها با خطا مواجه شدند.")

rotating_llm = RotatingGeminiLLM(API_KEYS)

def safe_generate_content(*, model, contents, config):
    for api_key in API_KEYS:
        try:
            client = genai.Client(api_key=api_key)
            response = client.models.generate_content(
                model=model,
                contents=contents,
                config=config
            )
            return response
        except Exception as e:
            print(f"⚠️ خطا با API {api_key[:10]}...: {str(e)}")
            continue
    raise RuntimeError("❌ تمام API Keyها با خطا مواجه شدند.")

llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash", google_api_key="AIzaSyC8tN4kY2QU5ACRacPazzRQeJPtAC08Vm8")

RESULT_FILE_PATH = Path("resume_results.xlsx")
if RESULT_FILE_PATH.exists():
    RESULT_FILE_PATH.unlink()

os.environ['SSL_CERT_FILE'] = certifi.where()

proxy_url = "http://172.16.217.234:33525"
os.environ['HTTP_PROXY'] = proxy_url
os.environ['HTTPS_PROXY'] = proxy_url

test_url = "https://generativelanguage.googleapis.com/v1beta/models"
try:
    response = requests.get(test_url, proxies={"http": proxy_url, "https": proxy_url}, timeout=5)
    if response.status_code == 200:
        print("✅ اتصال برقرار است.")
    else:
        print(f"⚠️ کد وضعیت: {response.status_code}")
except Exception as e:
    print(f"❌ خطا در اتصال پراکسی: {e}")

pd.set_option('display.max_rows', None)
OUTPUT_ALL_PATH = Path("recruitment_score.xlsx")
BATCH_SIZE = 10
RESULT_FILE_PATH = Path("resume_results.xlsx")

if 'live_results' not in st.session_state:
    st.session_state['live_results'] = []

JOB_PROFILES = [
    {
        "id": "job_rnd_01",
        "title": "تحقیق و توسعه سامانه‌ها",
        "tasks": [
            "تحلیل و احصا نیازمندی‌های نرم‌افزاری ذینفعان در حوزه زیرساخت‌های پردازش حجیم، لاگ و گردش کار",
            "توسعه و پیاده‌سازی راهکارهای نرم‌افزاری شامل ذخیره‌سازی، گردش کار و پرتال",
            "اجرای فرآیند استقرار زیرساخت و سامانه‌ها",
            "پشتیبانی و نگهداشت سامانه‌ها و پاسخ به تیکت‌ها",
            "مستندسازی و مدیریت دانش سامانه‌ها"
        ],
        "competencies_technical": [
            {"name": "برنامه‌نویسی و مبانی علم کامپیوتر"},
            {"name": "تحلیل نیازمندی نرم‌افزار"},
            {"name": "زبان‌های برنامه‌نویسی بک‌اند یا فرانت (مثل Python یا JavaScript)"},
            {"name": "پایگاه داده"}
        ],
        "majors": ["مهندسی کامپیوتر", "مهندسی صنایع", "رشته‌های فنی و مهندسی"]
    },
    {
        "id": "job_spatial_01",
        "title": "توسعه راهکارهای تحلیل اطلاعات مکانی",
        "tasks": [
            "تحلیل نیازمندی‌های داده‌محور مکانی",
            "توسعه راهکارهای نرم‌افزاری GIS و RS",
            "فرایند ETL داده‌های مکانی",
            "استقرار و پشتیبانی راهکارهای داده‌محور GIS/RS",
            "مستندسازی پروژه‌ها و ماموریت‌های مکانی"
        ],
        "competencies_technical": [
            {"name": "مبانی سنجش از دور"},
            {"name": "ابزارهای RS مانند ENVI، ERDAS، SNAP"},
            {"name": "مبانی هوش مصنوعی / پردازش تصویر"},
            {"name": "برنامه‌نویسی Python / MATLAB"},
            {"name": "نرم‌افزارهای GIS مانند ArcGIS/QGIS"}
        ],
        "majors": ["نقشه‌برداری", "مهندسی کامپیوتر", "مهندسی برق"]
    },
    {
        "id": "job_ai_01",
        "title": "توسعه راهکارهای مبتنی بر هوش مصنوعی",
        "tasks": [
            "تحلیل نیازهای داده‌محور با تاکید بر AI",
            "پیاده‌سازی مدل‌های آماری و یادگیری ماشین",
            "استقرار مدل‌ها با ابزارهای MLOps",
            "تهیه گزارشات تحلیلی",
            "مدیریت دانش پروژه‌های AI"
        ],
        "competencies_technical": [
            {"name": "مدل‌سازی آماری / یادگیری ماشین"},
            {"name": "برنامه‌نویسی Python / R / GAMS"},
            {"name": "کار با پایگاه داده"}
        ],
        "majors": ["علوم کامپیوتر", "ریاضی", "آمار", "مهندسی صنایع", "اقتصاد", "مهندسی مالی", "برق"]
    },
    {
        "id": "job_research_01",
        "title": "کارشناس ارتباط با مراکز پژوهشی",
        "tasks": [
            "احصا مسائل فناورانه و داده‌محور",
            "پشتیبانی سامانه دانش نظام مسائل",
            "مطالعات تطبیقی در حوزه هوش مصنوعی",
            "مستندسازی اسناد راهبردی AI",
            "رصد و تحلیل فناوری‌های نوظهور"
        ],
        "competencies_technical": [
            {"name": "Microsoft Office"},
            {"name": "مبانی علم داده و IT"},
            {"name": "اصول تحقیق و توسعه"}
        ],
        "majors": ["مدیریت", "مهندسی صنایع", "علوم اقتصادی", "علوم کامپیوتر"]
    },
    {
        "id": "job_analysis_01",
        "title": "کارشناس تحلیلگر داده و هوش تجاری",
        "tasks": [
            "گروه بندی و مرتب کردن اطلاعات",
            "تحلیل داده های مربوط به کسب و کار",
            "تمیزسازی داده ها ETL",
            "مستندسازی اسناد راهبردی ",
            "نامه نگاری و مکاتبات اداری",
            "بصری سازی داده ها"
        ],
        "competencies_technical": [
            {"name": "Microsoft Office"},
            {"name": "شناخت و تحلیل کسب و کار"},
            {"name": "ابزارهای مصورسازی مانند powerBI"},
            {"name": "ابزارهای تحلیل داده مانند KNIME"},
            {"name": "آشنایی با زان های برنامه نویسی مانند python , R"}
        ],
        "majors": ["مدیریت", "مهندسی صنایع", "علوم اقتصادی", "مهندسی کامپیوتر"]
    }
]

universities_info = [
    "دانشگاه تهران (برتر، دولتی)",
    "دانشگاه صنعتی شریف (برتر، دولتی)",
    "دانشگاه صنعتی امیرکبیر (برتر، دولتی)",
    "دانشگاه علم و صنعت ایران (برتر، دولتی)",
    "دانشگاه خواجه نصیر (برتر، دولتی)",
    "دانشگاه خوارزمی (برتر، دولتی)",
    "دانشگاه فردوسی مشهد (دولتی)",
    "دانشگاه تبریز (دولتی)",
    "دانشگاه اصفهان (دولتی)",
    "دانشگاه صنعتی اصفهان (دولتی)",
    "دانشگاه آزاد اسلامی (آزاد)",
    "دانشگاه پیام نور (پیام نور)",
    "دانشگاه غیرانتفاعی (غیرانتفاعی)",
    "دانشگاه علمی کاربردی (علمی کاربردی)"
]

AGENT_WEIGHTS = {
    "SkillAgent": 0.40,
    "ExperienceAgent": 0.30,
    "EducationAgent": 0.20,
    "VolunteeringAgent": 0.05,
    "SoftSkillsAgent": 0.05
}

def score_text_section(text): 
    if not text or str(text).strip() == "": 
        return 30

    prompt = f"""  
    Please rate the quality of the following resume section on a scale of 0 to 100.  
    Consider clarity, relevance, and value in a resume.  
    Return only a number between 0.0 and 1.0.  

    Text: 
    \"\"\" 
    {text} 
    \"\"\" 
    """ 

    try: 
        response = llm.invoke([HumanMessage(content=prompt)]) 
        score = float(response.content.strip()) 
        return round(max(0.0, min(1.0, score)) * 100, 2)
    except: 
        return 30

def process_batch(batch_df, prompt_text):
    payload = {
        "employer requirements": prompt_text,
        "applicant information": [
            {"resume": " ".join([str(row[col]) for col in batch_df.columns]), "id": str(idx)}
            for idx, row in batch_df.iterrows()
        ]
    }
    try:
        response = safe_generate_content(
            model='gemini-2.5-flash',
            contents=json.dumps(payload, ensure_ascii=False),
            config={
                'response_mime_type': 'application/json',
                'system_instruction': """
شما یک ارزیاب حرفه‌ای منابع انسانی هستید. معیارهای ارزیابی:
- تطابق مهارت‌های نرم‌افزاری
- تطابق سوابق شغلی
- مقطع و رشته تحصیلی مرتبط
- دانشگاه دولتی و معتبر
- سن مناسب (۲۲ تا ۳۵)
- حقوق درخواستی (۲۰ تا ۴۵ میلیون)
امتیاز بین ۱ تا ۱۰ بدهید. اگر اطلاعات نبود، بنویسید: 'اطلاعات کافی نیست'.
""",
                'response_schema': {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "score": {"type": "number", "nullable": False},
                            "check_id": {"type": "string", "nullable": False},
                            "why": {"type": "string", "nullable": False}
                        }
                    }
                },
                'temperature': 0
            }
        )
        result = json.loads(response.candidates[0].content.parts[0].text)
        return pd.DataFrame(result)
    except Exception:
        return pd.DataFrame([{
            "score": 1.0,
            "check_id": str(idx),
            "why": "خطا در پردازش - اطلاعات کافی نیست"
        } for idx, row in batch_df.iterrows()])

def to_excel(df, path):
    df.to_excel(path, index=False)

def match_resume_to_job_parallel(resume_text, job_profiles, threshold=7):
    best_match = None
    best_score = -1
    best_reason = ""
    log_messages = []

    def evaluate_job_with_key(api_key, job):
        prompt = f"""بر اساس اطلاعات زیر:
رزومه:
{resume_text}

موقعیت شغلی:
عنوان: {job['title']}
شرح وظایف: {'؛ '.join(job['tasks'])}
مهارت‌های تخصصی: {'؛ '.join([c['name'] for c in job.get('competencies_technical', [])])}
رشته‌های مرتبط: {'؛ '.join(job.get('majors', []))}

آیا این رزومه با این موقعیت شغلی تطابق دارد؟ لطفاً:
- یک امتیاز بین ۰ تا 100 بده
- در صورت مناسب بودن، دلیل را شرح بده
- در صورت نامناسب بودن، بنویس چرا مناسب نیست

لطفاً همیشه پاسخ را به فرمت زیر و با هردو بخش بده:
امتیاز: [یک عدد از 0 تا 100]
دلیل: [یک جمله واضح و دقیق شامل دلیل انتخاب یا عدم انتخاب]

"""
        try:
            response = safe_generate_content_for_key(
                api_key=api_key,
                model="gemini-2.5-flash",
                contents=prompt,
                config={"temperature": 0}
            )
            if isinstance(response, dict) and "error" in response:
                return None

            text = response.candidates[0].content.parts[0].text.strip()
            lines = [line.strip() for line in text.splitlines() if line.strip() != ""]

            score = -1
            reason = "توضیحی ارائه نشده است"

            for line in lines:
                if line.startswith("امتیاز"):
                    try:
                        score = int("".join(filter(str.isdigit, line)))
                    except:
                        score = -1
                if line.startswith("دلیل"):
                    reason = line.replace("دلیل:", "").strip()

            if reason == "توضیحی ارائه نشده است":
                for i, line in enumerate(lines):
                    if "امتیاز" in line and i + 1 < len(lines):
                        possible_reason = lines[i + 1]
                        if not possible_reason.startswith("امتیاز") and "دلیل" not in possible_reason:
                            reason = possible_reason
                            break

            return {"title": job["title"], "score": score, "reason": reason}

        except Exception as e:
            return None

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_job = {
            executor.submit(evaluate_job_with_key, api_key, job): job
            for api_key, job in zip(API_KEYS * (len(job_profiles) // len(API_KEYS) + 1), job_profiles)
        }

        for future in concurrent.futures.as_completed(future_to_job):
            result = future.result()
            if result:
                log_messages.append(f"🔹 {result['title']} → امتیاز: {result['score']} | دلیل: {result['reason']}")
                if result["score"] > best_score:
                    best_score = result["score"]
                    best_match = result["title"]
                    best_reason = result["reason"]

    log = "\n".join(log_messages)

    if best_score >= threshold:
        return best_match, best_reason, log
    else:
        return "مناسب هیچکدام از شناسنامه‌های شغلی نمی‌باشد", best_reason or "رزومه تطابق کافی با هیچ‌کدام از شغل‌ها ندارد.", log


def apply_matching_to_batch(batch_df):
    all_results = []

    for idx, row in batch_df.iterrows():
        resume_text = " ".join([str(row[col]) for col in batch_df.columns])
        match_df = evaluate_resume_against_all_jobs(resume_text, JOB_PROFILES)

        match_df["ردیف رزومه"] = idx + 1
        match_df["نام"] = row.get("نام", "")
        match_df["نام خانوادگی"] = row.get("نام خانوادگی", "")

        all_results.append(match_df)

    final_df = pd.concat(all_results, ignore_index=True)
    return final_df

top_universities = ['دانشگاه صنعتی شریف', 'دانشگاه تهران', 'دانشگاه صنعتی امیرکبیر', 'دانشگاه علم و صنعت ایران']
public_keywords = ['صنعتی', 'تهران', 'امیرکبیر', 'علم و صنعت', 'فردوسی', 'تبریز', 'اصفهان', 'دولتی']

def is_public_university(univ_name):
    return any(keyword in str(univ_name) for keyword in public_keywords)

def is_top_university(univ_name):
    return any(top in str(univ_name) for top in top_universities)

def color_score_column(val):
    if val >= 9:
        color = '#00C853'
    elif val >= 8:
        color = '#AEEA00'
    elif val >= 7:
        color = '#FFD600'
    elif val >= 6:
        color = '#FF9100'
    elif val >= 5:
        color = '#FF3D00'
    else:
        color = '#D50000'
    return f'background-color: {color}; color: white; font-weight: bold'


def adjust_score(row):
    score = row['score']
    if 'سن' in row and (row['سن'] < 22 or row['سن'] > 35):
        score -= 1
    if 'حقوق درخواستی' in row and (row['حقوق درخواستی'] < 20 or row['حقوق درخواستی'] > 45):
        score -= 1
    if 'مقطع تحصیلی' in row and 'کارشناسی' not in str(row['مقطع تحصیلی']):
        score -= 0.5
    univ = row.get('نام دانشگاه', '')
    if is_public_university(univ):
        score += 0.5
    if is_top_university(univ):
        score += 0.5
    return max(min(score, 10), 1.0)

def skill_agent(resume, skills):
    prompt = f"""
    شما یک ارزیاب منابع انسانی هستید. فقط مهارت‌های زیر را در رزومه زیر بررسی کن:
    مهارت‌های مورد انتظار: {', '.join(skills)}
    رزومه:
    {resume}
    
    یک عدد بین ۰ تا ۱۰۰ به میزان تطابق مهارت‌های رزومه با مهارت‌های مورد انتظار بده و یک جمله دلیل برای این امتیاز بنویس.
    فرمت دقیق پاسخ:
    امتیاز: [یک عدد]
    دلیل: [یک جمله کوتاه]
    """
    messages = [HumanMessage(content=prompt)]
    result = rotating_llm.invoke(messages)
    text = result.content
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    score, reason = 0, ""
    for line in lines:
        if line.startswith("امتیاز"):
            score = int("".join(filter(str.isdigit, line)))
        if line.startswith("دلیل"):
            reason = line.replace("دلیل:", "").strip()
    return score, reason

skill_tool = Tool(
    name="SkillAgent",
    func=lambda input: skill_agent(input["resume"], input["skills"]),
    description="ارزیابی مهارت‌های نرم‌افزاری"
)
experience_tool = Tool(
    name="ExperienceAgent",
    func=lambda input: experience_agent(input["resume"], input["required_experience_desc"]),
    description="ارزیابی تجربه شغلی"
)
education_tool = Tool(
    name="EducationAgent",
    func=lambda input: education_agent(
        input["resume"],
        input["university_list"],
        input["major_list"],
        input["job_profile_title"]
    ),
    description="ارزیابی تحصیلات دانشگاهی"
)
volunteering_tool = Tool(
    name="VolunteeringAgent",
    func=lambda input: volunteering_agent(input["resume"], input.get("volunteering_field")),
    description="ارزیابی فعالیت داوطلبانه"
)
softskills_tool = Tool(
    name="SoftSkillsAgent",
    func=lambda input: softskills_agent(input["resume"], input.get("about_me_field")),
    description="ارزیابی مهارت‌های نرم و شایستگی فردی"
)

def experience_agent(resume, required_experience_desc):
    prompt = f"""
    شما یک ارزیاب منابع انسانی هستید. فقط تجربه‌های شغلی رزومه زیر را از نظر میزان تطابق با نیازهای شغلی زیر بررسی کن:
    سابقه مورد انتظار: {required_experience_desc}
    رزومه:
    {resume}
    
    یک عدد بین ۰ تا ۱۰۰ به میزان تطابق سابقه کاری رزومه با نیازهای موقعیت شغلی بده و یک جمله دلیل برای این امتیاز بنویس.
    فرمت دقیق پاسخ:
    امتیاز: [یک عدد]
    دلیل: [یک جمله کوتاه]
    """
    messages = [HumanMessage(content=prompt)]
    result = rotating_llm.invoke(messages)
    text = result.content
    lines = [l.strip() for l in text.splitlines() if l.strip()]  
    score, reason = 0, ""
    for line in lines:
        if line.startswith("امتیاز"):
            score = int("".join(filter(str.isdigit, line)))
        if line.startswith("دلیل"):
            reason = line.replace("دلیل:", "").strip()
    return score, reason


def education_agent(resume, universities_info, major_list, job_profile_title):
    prompt = f"""
    شما یک ارزیاب منابع انسانی هستید. بخش تحصیلات رزومه زیر را فقط از نظر سه معیار بررسی کن:
    ۱. اعتبار دانشگاه و نوع آن (در فهرست زیر دانشگاه‌های معتبر و نوع هرکدام آمده است، دانشگاه‌های برتر و دولتی امتیاز بیشتری دارند، آزاد و پیام نور امتیاز متوسط، غیرانتفاعی و علمی کاربردی امتیاز پایین‌تر):
    {chr(10).join(universities_info)}
    ۲. تطابق رشته تحصیلی با موقعیت شغلی "{job_profile_title}" (لیست رشته‌های مطلوب: {', '.join(major_list)})
    ۳. مدت زمان تحصیل کارشناسی (زیر ۴ سال عالی، ۴ سال خوب، بیشتر از ۴ سال ضعیف)
    
    رزومه:
    {resume}

    یک عدد بین ۰ تا ۱۰۰ به میزان شایستگی تحصیلات رزومه نسبت به نیازهای شغلی بده و یک جمله دلیل برای این امتیاز بنویس.
    فرمت دقیق پاسخ:
    امتیاز: [یک عدد]
    دلیل: [یک جمله کوتاه]
    """
    messages = [HumanMessage(content=prompt)]
    result = rotating_llm.invoke(messages)
    text = result.content
    lines = [l.strip() for l in text.splitlines() if l.strip()]  
    score, reason = 0, ""
    for line in lines:
        if line.startswith("امتیاز"):
            score = int("".join(filter(str.isdigit, line)))
        if line.startswith("دلیل"):
            reason = line.replace("دلیل:", "").strip()
    return score, reason

def volunteering_agent(resume, volunteering_field=None):
    field = volunteering_field if volunteering_field else resume
    prompt = f"""
    شما یک ارزیاب منابع انسانی هستید. فقط فعالیت‌های داوطلبانه و کارهای اجتماعی رزومه زیر را بررسی کن:
    اگر فعالیت داوطلبانه مرتبط و تأثیرگذار (در سطح بالا) باشد، امتیاز بالا بده، اگر نباشد یا کم باشد امتیاز پایین.
    رزومه/فعالیت داوطلبانه:
    {field}

    یک عدد بین ۰ تا ۱۰۰ به فعالیت داوطلبانه بده و یک جمله دلیل برای این امتیاز بنویس.
    فرمت دقیق پاسخ:
    امتیاز: [یک عدد]
    دلیل: [یک جمله کوتاه]
    """
    messages = [HumanMessage(content=prompt)]
    result = rotating_llm.invoke(messages)
    text = result.content
    lines = [l.strip() for l in text.splitlines() if l.strip()]  
    score, reason = 0, ""
    for line in lines:
        if line.startswith("امتیاز"):
            score = int("".join(filter(str.isdigit, line)))
        if line.startswith("دلیل"):
            reason = line.replace("دلیل:", "").strip()
    return score, reason

def softskills_agent(resume, about_me_field=None):
    field = about_me_field if about_me_field else resume
    prompt = f"""
    شما یک ارزیاب منابع انسانی هستید. فقط مهارت‌های نرم و شایستگی‌های فردی رزومه زیر را بررسی کن:
    ویژگی‌هایی مثل: کار تیمی، ارتباط موثر، مدیریت، مسئولیت‌پذیری، دقت، میل به یادگیری و هوش هیجانی (EQ) را تحلیل کن.
    اگر رزومه یا بخش 'درباره من' شواهد قوی از این ویژگی‌ها دارد امتیاز بالا بده، اگر نداشت یا ضعیف بود امتیاز پایین.
    متن برای تحلیل:
    {field}

    یک عدد بین ۰ تا ۱۰۰ به مهارت‌های نرم بده و یک جمله دلیل برای این امتیاز بنویس.
    فرمت دقیق پاسخ:
    امتیاز: [یک عدد]
    دلیل: [یک جمله کوتاه]
    """
    messages = [HumanMessage(content=prompt)]
    result = rotating_llm.invoke(messages)
    text = result.content
    lines = [l.strip() for l in text.splitlines() if l.strip()]  
    score, reason = 0, ""
    for line in lines:
        if line.startswith("امتیاز"):
            score = int("".join(filter(str.isdigit, line)))
        if line.startswith("دلیل"):
            reason = line.replace("دلیل:", "").strip()
    return score, reason

def scoring_chain(
    resume,
    skills,
    required_experience_desc,
    universities_info,
    major_list,
    job_profile_title,
    volunteering_field=None,
    about_me_field=None
):
    results = {}

    skill_score, skill_reason = skill_agent(resume, skills)
    results["SkillAgent"] = {"score": skill_score, "reason": skill_reason}

    exp_score, exp_reason = experience_agent(resume, required_experience_desc)
    results["ExperienceAgent"] = {"score": exp_score, "reason": exp_reason}

    edu_score, edu_reason = education_agent(resume, universities_info, major_list, job_profile_title)
    results["EducationAgent"] = {"score": edu_score, "reason": edu_reason}

    vol_score, vol_reason = volunteering_agent(resume, volunteering_field)
    results["VolunteeringAgent"] = {"score": vol_score, "reason": vol_reason}

    soft_score, soft_reason = softskills_agent(resume, about_me_field)
    results["SoftSkillsAgent"] = {"score": soft_score, "reason": soft_reason}

    results["VolunteeringAgent"]["score"] = score_text_section(vol_reason)
    results["SoftSkillsAgent"]["score"] = score_text_section(soft_reason)

    final_score = 0
    for agent, w in AGENT_WEIGHTS.items():
        final_score += results[agent]["score"] * w

    final_score = round(final_score / sum(AGENT_WEIGHTS.values()), 2)
    results["FinalScore"] = final_score

    return results


def evaluate_resume_against_all_jobs(resume_text, job_profiles):
    prompt = f"""شما یک ارزیاب منابع انسانی هستید. با توجه به رزومه زیر، لطفاً برای هر یک از موقعیت‌های شغلی تعریف‌شده، یک درصد تطابق بین ۰ تا ۱۰۰ بدهید و یک دلیل منطقی برای آن ذکر کنید.

رزومه:
{resume_text}

ساختار پاسخ دقیقا به صورت JSON زیر باشد:
[
  {{
    "title": "عنوان شغل اول",
    "match_percent": 85,
    "reason": "توضیح دلیل تطابق یا عدم تطابق"
  }},
  {{
    "title": "عنوان شغل دوم",
    "match_percent": 45,
    "reason": "..."
  }}
  ...
]
موقعیت‌های شغلی:
{json.dumps(job_profiles, ensure_ascii=False)}
"""

    try:
        response = safe_generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config={
                "response_mime_type": "application/json",
                "temperature": 0
            }
        )
        json_text = response.candidates[0].content.parts[0].text.strip()
        parsed = json.loads(json_text)
        return pd.DataFrame(parsed)
    except Exception as e:
        st.error(f"❌ خطا در تحلیل تطابق: {e}")
        return pd.DataFrame()

def process_resume_row(row, row_index):
    resume_text = " ".join([str(row[col]) for col in row.index])
    title, reason, log = match_resume_to_job(resume_text, JOB_PROFILES)

    gemini_df = process_batch(pd.DataFrame([row]), prompt_text="ارزیابی عمومی رزومه")
    initial_score = gemini_df.iloc[0]['score']

    score = adjust_score({**row.to_dict(), 'score': initial_score})

    new_data = row.to_dict()
    new_data.update({
        "ردیف": row_index + 1,
        "score": score,
        "دلیل": gemini_df.iloc[0]['why'],
        "موقعیت شغلی پیشنهادی": title,
        "دلیل انتخاب موقعیت شغلی": reason,
        "گزارش بررسی شناسنامه‌ها": log
    })

    if RESULT_FILE_PATH.exists():
        existing = pd.read_excel(RESULT_FILE_PATH)
        updated = pd.concat([existing, pd.DataFrame([new_data])], ignore_index=True)
    else:
        updated = pd.DataFrame([new_data])

    updated.to_excel(RESULT_FILE_PATH, index=False)

    st.session_state['live_results'].append(new_data)
    return new_data

st.markdown("""
    <style>
    .custom-title {
        font-size: 50px !important;
        color: #1a73e8 !important;
        font-weight: bold !important;
        text-align: center !important;
        margin-top: 40px !important;
        margin-bottom: 30px !important;
    }
    </style>
""", unsafe_allow_html=True)
st.markdown('<div class="custom-title">📋 سامانه هوشمند ارزیابی رزومه</div>', unsafe_allow_html=True)
st.markdown("<p style='font-size: 16px; color: #555;'>ارزیابی هوشمند رزومه‌ها بر اساس معیارهای منابع انسانی، شناسنامه‌های شغلی و مهارت‌های تخصصی.</p>", unsafe_allow_html=True)

uploaded_file = st.file_uploader("📄 فایل اکسل رزومه‌ها را بارگذاری کنید:", type=["xlsx"])

with st.sidebar:
    st.markdown("## 📊 وضعیت سیستم")
    st.markdown("### ⏳ پردازش رزومه‌ها")
    status_placeholder = st.empty()
    progress_placeholder = st.empty()

if uploaded_file and ('live_results' not in st.session_state or len(st.session_state['live_results']) == 0):
    status_placeholder.info("✅ فایل آپلود شده. آماده برای شروع ارزیابی...")
    progress_placeholder.progress(0)
elif not uploaded_file:
    status_placeholder.info("⏳ منتظر آپلود فایل رزومه باشید.")
    progress_placeholder.progress(0)

with st.sidebar:
    if st.button("🔄 ریست کامل اطلاعات"):
        for key in ['final_df', 'live_results']:
            if key in st.session_state:
                del st.session_state[key]
        if RESULT_FILE_PATH.exists():
            RESULT_FILE_PATH.unlink()
        st.success("✅ اطلاعات با موفقیت ریست شد.")

job_titles = [job['title'] for job in JOB_PROFILES]

selected_job_titles = st.multiselect(
    "عنوان شغلی مورد نظر را انتخاب کنید (امکان انتخاب چندتایی):",
    options=job_titles,
    default=None
)

custom_job_title = st.text_input("در صورتی که عنوان شغلی مورد نظر شما در لیست نبود، اینجا وارد کنید:")

all_selected_titles = selected_job_titles.copy()
if custom_job_title.strip() != "":
    all_selected_titles.append(custom_job_title.strip())

selected_skills = []
for job in JOB_PROFILES:
    if job["title"] in all_selected_titles:
        selected_skills.extend([c['name'] for c in job.get('competencies_technical', [])])

selected_skills = list(sorted(set(selected_skills)))

edited_skills = st.multiselect(
    "مهارت‌های مورد نیاز را بررسی و ویرایش کنید:",
    options=selected_skills,
    default=selected_skills
)

custom_skill = st.text_input("در صورت نیاز، مهارت جدید وارد کنید:")

all_skills = edited_skills.copy()
if custom_skill.strip() and custom_skill.strip() not in all_skills:
    all_skills.append(custom_skill.strip())

def process_single_resume(args):
    """Process a single resume with a specific API key"""
    idx, row, api_key, all_skills = args
    start_time = time.time()
    
    try:
        llm_instance = ChatGoogleGenerativeAI(model="gemini-2.5-flash", google_api_key=api_key)
        
        resume = " ".join([str(row[col]) for col in row.index]) 
        required_experience_desc = "سابقه مرتبط با عنوان شغلی" 
        universities = universities_info 
        major_list = []
        job_profile_title = ""
        volunteering_field = row.get("فعالیت داوطلبانه", "") 
        about_me_field = row.get("درباره من", "")

        results = scoring_chain(
            resume, 
            all_skills, 
            required_experience_desc, 
            universities, 
            major_list, 
            job_profile_title, 
            volunteering_field, 
            about_me_field
        )

        row_data = row.to_dict()
        row_data['ردیف'] = idx + 1
        for agent, detail in results.items():
            if agent != "FinalScore":
                row_data[f"{agent}_score"] = detail['score']
                row_data[f"{agent}_reason"] = detail['reason']
        row_data['final_score'] = results['FinalScore']
        row_data['تایید و رد اولیه'] = "تایید" if row_data['final_score'] >= 70 else "رد"

        processing_time = round(time.time() - start_time, 2)
        row_data['زمان پردازش (ثانیه)'] = processing_time 
        
        return (idx, row_data, None)
    
    except Exception as e:
        processing_time = round(time.time() - start_time, 2)
        return (idx, None, str(e))

if uploaded_file:
    df = pd.read_excel(uploaded_file, skiprows=2)
    
    st.info(f"تعداد رزومه‌های بارگذاری شده: {len(df)} | تعداد ستون‌ها: {len(df.columns)}")
    
    with st.expander("نمایش پیش‌نمایش داده‌ها"):
        st.dataframe(df.head())
    
    stage = st.radio("🧩 مرحله موردنظر را انتخاب کنید:", ["امتیازدهی", "تطبیق با شناسنامه‌های شغلی"])

    if stage == "امتیازدهی": 
        st.markdown("### 🚀 مرحله امتیازدهی رزومه‌ها") 
        
        max_workers = min(len(API_KEYS), len(df))
        st.info(f"پردازش موازی با {max_workers} API Key برای {len(df)} رزومه")
        
        if st.button("شروع امتیازدهی"): 
            total_start_time = time.time()
            results_placeholder = st.empty() 
            progress_bar = st.progress(0) 
            rows = [None] * len(df)
            completed = 0
            
            processing_args = [
                (idx, row, API_KEYS[idx % len(API_KEYS)], all_skills)
                for idx, (_, row) in enumerate(df.iterrows())
            ]
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_idx = {
                    executor.submit(process_single_resume, args): args[0] 
                    for args in processing_args
                }
                
                for future in concurrent.futures.as_completed(future_to_idx):
                    idx, row_data, error = future.result()
                    
                    if error:
                        st.warning(f"⚠️ خطا در پردازش رزومه ردیف {idx + 1}: {error}")
                        row_data = df.iloc[idx].to_dict()
                        row_data['ردیف'] = idx + 1
                        row_data['تایید و رد اولیه'] = "خطا"
                        row_data['final_score'] = 0
                    
                    rows[idx] = row_data
                    completed += 1
                    
                    progress_bar.progress(completed / len(df))
                    
                    current_results = [r for r in rows if r is not None]
                    if current_results:
                        temp_df = pd.DataFrame(current_results)
                        results_placeholder.dataframe(temp_df)
                    
                    live_df = pd.DataFrame(current_results)
                    total = len(df)
                    checked = len(live_df)
                    accepted = (live_df['تایید و رد اولیه'] == 'تایید').sum() if 'تایید و رد اولیه' in live_df.columns else 0
                    failed = (live_df['تایید و رد اولیه'] != 'تایید').sum() if 'تایید و رد اولیه' in live_df.columns else 0
                    
                     if 'زمان پردازش (ثانیه)' in live_df.columns:
                        avg_time = live_df['زمان پردازش (ثانیه)'].mean()
                        estimated_remaining = avg_time * (total - checked)
                        status_placeholder.markdown(f"""
                        **بررسی شده: {checked} / {total}**  
                        🟢 قبول‌شده: {accepted}  
                        🔴 رد‌شده: {failed}  
                        ⏱️ میانگین زمان هر رزومه: {avg_time:.2f}s  
                        ⏳ تخمین زمان باقیمانده: {estimated_remaining:.1f}s ({estimated_remaining/60:.1f} دقیقه)
                        """)
                    else:
                        status_placeholder.success(f"بررسی شده: {checked} / {total}")
                        status_placeholder.markdown(f"🟢 قبول‌شده: {accepted}")
                        status_placeholder.markdown(f"🔴 رد‌شده: {failed}")
                    progress_placeholder.progress(checked / total)
            
            results_df = pd.DataFrame(rows)
            results_placeholder.dataframe(results_df)
            results_df.to_excel("resume_scoring.xlsx", index=False)
            style_excel("resume_scoring.xlsx")

            total_time = round(time.time() - total_start_time, 2)  
            st.success(f"✅ امتیازدهی به پایان رسید. زمان کل: {total_time} ثانیه ({total_time/60:.2f} دقیقه)")

            with open("resume_scoring.xlsx", "rb") as f:
                st.download_button(
                    label="📥 دانلود فایل اکسل امتیازدهی",
                    data=f,
                    file_name="resume_scoring.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    elif stage == "تطبیق با شناسنامه‌های شغلی":
        st.markdown("### 🔍 مرحله تطبیق با شناسنامه‌های شغلی")
        results_placeholder = st.empty()
        progress_bar = st.progress(0)
        
        max_workers = min(len(API_KEYS), len(df))
        st.info(f"پردازش موازی با {max_workers} API Key برای {len(df)} رزومه")

        if st.button("🚀 شروع تطبیق با شناسنامه‌های شغلی"):
            total_start_time = time.time() 
            try:
                def process_single_matching(args):
                    """Process job matching for a single resume"""
                    idx, row, api_key = args
                    start_time = time.time()
                    try:
                        resume_text = " ".join([str(row[col]) for col in row.index])
                        
                        prompt = f"""شما یک ارزیاب منابع انسانی هستید. با توجه به رزومه زیر، لطفاً برای هر یک از موقعیت‌های شغلی تعریف‌شده، یک درصد تطابق بین ۰ تا ۱۰۰ بدهید و یک دلیل منطقی برای آن ذکر کنید.

رزومه:
{resume_text}

ساختار پاسخ دقیقا به صورت JSON زیر باشد:
[
  {{
    "title": "عنوان شغل اول",
    "match_percent": 85,
    "reason": "توضیح دلیل تطابق یا عدم تطابق"
  }},
  {{
    "title": "عنوان شغل دوم",
    "match_percent": 45,
    "reason": "..."
  }}
  ...
]
موقعیت‌های شغلی:
{json.dumps(JOB_PROFILES, ensure_ascii=False)}
"""
                        
                        client = genai.Client(api_key=api_key)
                        response = client.models.generate_content(
                            model="gemini-2.5-flash",
                            contents=prompt,
                            config={
                                "response_mime_type": "application/json",
                                "temperature": 0
                            }
                        )
                        
                        json_text = response.candidates[0].content.parts[0].text.strip()
                        parsed = json.loads(json_text)
                        match_df = pd.DataFrame(parsed)
                        
                        match_df["ردیف رزومه"] = idx + 1
                        match_df["نام"] = row.get("نام", "")
                        match_df["نام خانوادگی"] = row.get("نام خانوادگی", "")
                        processing_time = round(time.time() - start_time, 2)  # ADD THIS LINE
                        match_df["زمان پردازش (ثانیه)"] = processing_time

                        return (idx, match_df, None)
                    except Exception as e:
                        return (idx, None, str(e))
                
                processing_args = [
                    (idx, row, API_KEYS[idx % len(API_KEYS)])
                    for idx, (_, row) in enumerate(df.iterrows())
                ]
                
                all_results = [None] * len(df)
                completed = 0
                
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_idx = {
                        executor.submit(process_single_matching, args): args[0]
                        for args in processing_args
                    }
                    
                    for future in concurrent.futures.as_completed(future_to_idx):
                        idx, match_df, error = future.result()
                        
                        if error:
                            st.warning(f"⚠️ خطا در تطبیق رزومه ردیف {idx + 1}: {error}")
                        else:
                            all_results[idx] = match_df
                        
                        completed += 1
                        progress_bar.progress(completed / len(df))
                
                match_results = pd.concat([r for r in all_results if r is not None], ignore_index=True)
                
                def make_sentence(row):
                    return f"میزان انطباق با موقعیت شغلی {row['title']} {int(row['match_percent'])}٪ است، زیرا: {row['reason']}"

                grouped = match_results.groupby("ردیف رزومه")

                final_rows = []
                for resume_row_num, group in grouped:
                    name = group["نام"].iloc[0]
                    family = group["نام خانوادگی"].iloc[0]
                    sentences = [make_sentence(row) for _, row in group.iterrows()]
                    full_text = "  ".join(sentences)
                    best_row = group.loc[group["match_percent"].idxmax()]
                    best_title = best_row["title"]

                    final_rows.append({
                        "ردیف رزومه": resume_row_num,
                        "نام": name,
                        "نام خانوادگی": family,
                        "موقعیت شغلی پیشنهادی": best_title,
                        "تحلیل نهایی": full_text
                    })

                summary_df = pd.DataFrame(final_rows)

                summary_path = "job_matching_summary.xlsx"
                summary_df.to_excel(summary_path, index=False)
                style_excel(summary_path)
                total_time = round(time.time() - total_start_time, 2)  # ADD THIS LINE
                st.success(f"✅ تطبیق با شناسنامه‌های شغلی با موفقیت انجام شد. زمان کل: {total_time} ثانیه ({total_time/60:.2f} دقیقه)")
                st.dataframe(summary_df)

                with open(summary_path, "rb") as f:
                    st.download_button("📥 دانلود فایل نهایی تحلیل‌شده", f, file_name=summary_path)

                progress_bar.progress(1.0)
            
            except Exception as e:
                st.error(f"❌ خطا در انجام تطبیق: {e}")

if RESULT_FILE_PATH.exists():
    final_df = pd.read_excel(RESULT_FILE_PATH)

    display_df = final_df.copy()
    display_df.index = display_df.index + 1
    display_df.index.name = "ردیف"

    st.markdown("### ✅ جدول نهایی رزومه‌های بررسی‌شده")
    
    if 'score' in display_df.columns:
        styled_df = display_df.style.applymap(color_score_column, subset=['score'])
        st.dataframe(styled_df)
    else:
        st.dataframe(display_df)

    style_excel(RESULT_FILE_PATH)
    with open(RESULT_FILE_PATH, "rb") as f:
        st.download_button("📥 دانلود فایل نهایی", f, file_name="resume_results.xlsx")



