import streamlit as st
import pathlib
import fitz
import json
import re
import pandas as pd
import shutil
import os
from google import genai
from google.genai import types
import certifi
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import tempfile
import zipfile
from io import BytesIO
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import itertools

# صفحه کانفیگ
st.set_page_config(
    page_title="پردازشگر هوشمند رزومه",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* ================================
   فونت فارسی B Homa
   ================================ */
@import url('https://cdn.jsdelivr.net/gh/rastikerdar/vazir-font@v30.1.0/dist/font-face.css');

@font-face {
    font-family: 'B Homa';
    src: url('https://cdn.jsdelivr.net/gh/font-store/BHoma@master/BHoma.woff2') format('woff2');
    font-weight: normal;
    font-style: normal;
}

* {
    font-family: 'B Homa', 'Vazir', 'Tahoma', sans-serif !important;
}

/* ================================
   متغیرهای رنگی مدرن
   ================================ */
:root {
    --primary-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    --secondary-gradient: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
    --success-gradient: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
    --dark-bg: #1a1a2e;
    --card-bg: #ffffff;
    --text-primary: #2d3436;
    --text-secondary: #636e72;
    --border-color: #e1e8ed;
    --shadow-sm: 0 2px 8px rgba(0,0,0,0.08);
    --shadow-md: 0 4px 16px rgba(0,0,0,0.12);
    --shadow-lg: 0 8px 32px rgba(0,0,0,0.16);
}

/* ================================
   تنظیمات کلی RTL
   ================================ */
.main .block-container {
    direction: rtl;
    text-align: right;
    padding: 2rem 3rem;
    max-width: 1400px;
}

body {
    background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
}

/* ================================
   هدر اصلی
   ================================ */
.modern-header {
    background: var(--primary-gradient);
    color: white;
    padding: 3rem 2rem;
    border-radius: 24px;
    text-align: center;
    margin-bottom: 2rem;
    box-shadow: var(--shadow-lg);
    position: relative;
    overflow: hidden;
}

.modern-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -50%;
    width: 200%;
    height: 200%;
    background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
    animation: pulse 4s ease-in-out infinite;
}

@keyframes pulse {
    0%, 100% { transform: scale(1); opacity: 0.5; }
    50% { transform: scale(1.1); opacity: 0.8; }
}

.modern-header h1 {
    font-size: 2.8rem;
    font-weight: 700;
    margin: 0;
    position: relative;
    z-index: 1;
}

.modern-header p {
    font-size: 1.2rem;
    margin-top: 0.5rem;
    opacity: 0.9;
    position: relative;
    z-index: 1;
}

/* ================================
   کارت‌های مدرن
   ================================ */
.modern-card {
    background: var(--card-bg);
    border-radius: 20px;
    padding: 2rem;
    box-shadow: var(--shadow-md);
    transition: all 0.3s ease;
    border: 1px solid var(--border-color);
    margin-bottom: 1.5rem;
}

.modern-card:hover {
    transform: translateY(-4px);
    box-shadow: var(--shadow-lg);
}

.modern-card h1, .modern-card h2, .modern-card h3, 
.modern-card h4, .modern-card h5, .modern-card h6,
.modern-card p, .modern-card span, .modern-card div,
.modern-card label {
    color: var(--text-primary) !important;
}

/* ================================
   متریک‌های زیبا
   ================================ */
.metric-modern {
    background: white;
    border-radius: 16px;
    padding: 1.5rem;
    text-align: center;
    box-shadow: var(--shadow-sm);
    transition: all 0.3s ease;
    border: 2px solid transparent;
    position: relative;
    overflow: hidden;
}

.metric-modern::before {
    content: '';
    position: absolute;
    top: 0;
    right: 0;
    width: 100%;
    height: 4px;
    background: var(--primary-gradient);
}

.metric-modern:hover {
    border-color: #667eea;
    transform: translateY(-2px);
    box-shadow: var(--shadow-md);
}

.metric-modern h3 {
    font-size: 2.5rem;
    font-weight: 700;
    margin: 0.5rem 0;
    background: var(--primary-gradient);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

.metric-modern p {
    color: var(--text-secondary);
    font-size: 1rem;
    margin: 0;
}

/* ================================
   دکمه‌های مدرن
   ================================ */
.stButton > button {
    width: 100%;
    background: var(--primary-gradient) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.8rem 2rem !important;
    font-size: 1.1rem !important;
    font-weight: 600 !important;
    transition: all 0.3s ease !important;
    box-shadow: var(--shadow-sm) !important;
}

.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: var(--shadow-md) !important;
}

.stDownloadButton > button {
    background: var(--success-gradient) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.8rem 2rem !important;
    font-weight: 600 !important;
}

/* ================================
   ورودی‌های مدرن
   ================================ */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div > div {
    direction: rtl !important;
    text-align: right !important;
    border-radius: 12px !important;
    border: 2px solid var(--border-color) !important;
    padding: 0.8rem 1rem !important;
    transition: all 0.3s ease !important;
}

.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #667eea !important;
    box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1) !important;
}

/* ================================
   تب‌های مدرن
   ================================ */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
    background: white;
    border-radius: 16px;
    padding: 8px;
    box-shadow: var(--shadow-sm);
    direction: rtl;
}

.stTabs [data-baseweb="tab"] {
    border-radius: 12px !important;
    padding: 12px 24px !important;
    font-weight: 600 !important;
    transition: all 0.3s ease !important;
    color: #dc2626 !important;
}

.stTabs [aria-selected="true"] {
    background: var(--primary-gradient) !important;
    color: #dc2626 !important;
}

.stTabs button div {
    color: #dc2626 !important;
}

/* ================================
   آپلود فایل مدرن
   ================================ */
.stFileUploader {
    direction: rtl;
}

.stFileUploader > div {
    background: white;
    border: 3px dashed #667eea;
    border-radius: 20px;
    padding: 3rem 2rem;
    text-align: center;
    transition: all 0.3s ease;
}

.stFileUploader > div:hover {
    background: #f8f9ff;
    border-color: #764ba2;
}

.stFileUploader label,
.stFileUploader p,
.stFileUploader span,
.stFileUploader div {
    color: #2d3436 !important;
}

/* ================================
   پروگرس بار مدرن
   ================================ */
.stProgress > div > div > div {
    background: var(--primary-gradient) !important;
    border-radius: 10px !important;
    height: 12px !important;
}

/* ================================
   جدول داده‌ها - LTR برای نمایش درست
   ================================ */
.stDataFrame {
    direction: ltr !important;
    border-radius: 16px;
    overflow: hidden;
    box-shadow: var(--shadow-md);
}

.stDataFrame table {
    border-radius: 16px;
}

/* ================================
   اسلایدر - LTR برای عملکرد درست
   ================================ */
.stSlider {
    direction: ltr !important;
    padding: 1rem 0;
}

.stSlider > div > div > div > div {
    background: var(--primary-gradient) !important;
}

/* ================================
   باکس‌های اطلاعات
   ================================ */
.info-box-modern {
    background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
    border-right: 5px solid #2196f3;
    border-radius: 12px;
    padding: 1.5rem;
    margin: 1rem 0;
    direction: rtl;
}

.success-box-modern {
    background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);
    border-right: 5px solid #4caf50;
    border-radius: 12px;
    padding: 1.5rem;
    margin: 1rem 0;
    direction: rtl;
}

.error-box-modern {
    background: linear-gradient(135deg, #ffebee 0%, #ffcdd2 100%);
    border-right: 5px solid #f44336;
    border-radius: 12px;
    padding: 1.5rem;
    margin: 1rem 0;
    direction: rtl;
}

.warning-box-modern {
    background: linear-gradient(135deg, #fff3e0 0%, #ffe0b2 100%);
    border-right: 5px solid #ff9800;
    border-radius: 12px;
    padding: 1.5rem;
    margin: 1rem 0;
    direction: rtl;
}

/* ================================
   ساید‌بار
   ================================ */
.css-1d391kg, [data-testid="stSidebar"] {
    background: white !important;
    border-left: 1px solid var(--border-color);
}

.css-1d391kg .stSelectbox > div > div > div,
.css-1d391kg .stTextInput > div > div > input,
.css-1d391kg .stTextArea > div > div > textarea,
[data-testid="stSidebar"] .stSelectbox > div > div > div,
[data-testid="stSidebar"] .stTextInput > div > div > input,
[data-testid="stSidebar"] .stTextArea > div > div > textarea {
    direction: rtl !important;
    text-align: right !important;
}

/* متن‌های Sidebar همیشه مشکی */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4,
[data-testid="stSidebar"] h5,
[data-testid="stSidebar"] h6,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stMarkdown {
    color: #000000 !important;
}

[data-testid="stSidebar"] .stRadio label,
[data-testid="stSidebar"] .stCheckbox label {
    color: #000000 !important;
}

/* ================================
   Expander مدرن
   ================================ */
.streamlit-expanderHeader {
    background: white;
    border-radius: 12px;
    padding: 1rem 1.5rem;
    direction: rtl;
    font-weight: 600;
    border: 1px solid var(--border-color);
    transition: all 0.3s ease;
    color: var(--text-primary) !important;
}

.streamlit-expanderHeader:hover {
    background: #f8f9ff;
    border-color: #667eea;
}

.streamlit-expanderHeader p,
.streamlit-expanderHeader span,
.streamlit-expanderHeader div {
    color: var(--text-primary) !important;
}

/* ================================
   راست‌چین کردن متن‌ها
   ================================ */
h1, h2, h3, h4, h5, h6, p, div, span, label {
    direction: rtl !important;
    text-align: right !important;
}

/* رنگ پیش‌فرض متن‌ها */
.main p, .main div, .main span, .main label {
    color: var(--text-primary);
}

/* متن‌های داخل کانتینرهای سفید */
.stMarkdown, .element-container {
    color: var(--text-primary);
}

/* ================================
   انیمیشن‌های ظریف
   ================================ */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(10px); }
    to { opacity: 1; transform: translateY(0); }
}

.modern-card, .metric-modern {
    animation: fadeIn 0.5s ease-out;
}

/* ================================
   اسکرول‌بار سفارشی
   ================================ */
::-webkit-scrollbar {
    width: 10px;
    height: 10px;
}

::-webkit-scrollbar-track {
    background: #f1f1f1;
    border-radius: 10px;
}

::-webkit-scrollbar-thumb {
    background: var(--primary-gradient);
    border-radius: 10px;
}

::-webkit-scrollbar-thumb:hover {
    background: #5568d3;
}

/* ================================
   رادیو باتن و چک‌باکس
   ================================ */
.stRadio > div {
    direction: rtl;
}

.stCheckbox > label {
    direction: rtl;
}
</style>
""", unsafe_allow_html=True)

# تنظیمات اولیه
os.environ["SSL_CERT_FILE"] = certifi.where()

# کلیدهای API
DEFAULT_GENAI_KEYS = [
    "AIzaSyAQ1Z8HmIZm-eNvohxoM4ZNFM8JsZsxDII",  
    "AIzaSyAQhK01WbSxiXUdXqe5xEvJA3feUiQCL0E",  
    "AIzaSyAhMXCXIfat3NQqsyWk-S8gdOzTRZLc_bA",  
    "AIzaSyCBH-nSuALuLBerOBn2JS-z3yBYuvPXTPw",
    "AIzaSyClzhUwWrUyI_dEjaYO4d4mijfBFGw1his",
    "AIzaSyCWZVz-ciOp91vKr2u7J87IktK2skygOro",
    "AIzaSyB11u1-TTuvIRNhSAp44PgWWpoK9kq1mAo"
]

# الگوهای خطای محدودیت
_rate_limit_patterns = [
    re.compile(r"429"),
    re.compile(r"rate.*limit", re.IGNORECASE),
    re.compile(r"quota", re.IGNORECASE),
    re.compile(r"exceed", re.IGNORECASE),
    re.compile(r"RateLimit", re.IGNORECASE),
]

# ترتیب فیلدها
ORDERED_FIELDS = [
    "شناسه", "نام", "نام خانوادگی", "شماره تماس", "جنسیت", "ایمیل", "کانال دریافت رزومه",
    "معرف", "کارشناسی", "کارشناسی ارشد", "دکتری", "رشته تحصیلی", "گرایش تحصیلی", "مقطع تحصیلی",
    "دانشگاه محل تحصیلی", "نوع دانشگاه آخرین مدرک تحصیلی", "وضعیت تحصیلی",
    "دوره های آموزشی", "نرم افزارها", "سوابق شغلی",
    "وضعیت خدمت سربازی", "وضعیت تاهل", "محل سکونت", "سن", "year_of_birth",
    "حداقل حقوق ماهیانه", "حداکثر حقوق ماهیانه",
    "فعالیت های داوطلبانه", "درباره ی من",
    "تایید و رد اولیه", "علت رد"
]

class QuotaExhaustedException(Exception):
    """خطای سهمیه تمام شده برای همه کلیدها"""
    pass

class APIKeyManager:
    """مدیر کلیدهای API با قابلیت مدیریت موثر"""
    def __init__(self, api_keys):
        self.api_keys = api_keys
        self.clients = {}
        self.failed_keys = set()
        self.key_usage_count = {key: 0 for key in api_keys}
        self.lock = threading.Lock()
        self._initialize_clients()
    
    def _initialize_clients(self):
        """ایجاد کلاینت‌ها برای تمام کلیدها"""
        for key in self.api_keys:
            try:
                client = genai.Client(api_key=key)
                self.clients[key] = client
            except Exception as e:
                print(f"Failed to initialize client for key {key[:10]}...: {e}")
    
    def get_available_client(self):
        """دریافت یک کلاینت موجود"""
        with self.lock:
            available_keys = [key for key in self.api_keys if key not in self.failed_keys]
            
            if not available_keys:
                raise QuotaExhaustedException("همه کلیدهای API غیرفعال شده‌اند")
            
            selected_key = min(available_keys, key=lambda k: self.key_usage_count[k])
            self.key_usage_count[selected_key] += 1
            
            return self.clients[selected_key], selected_key
    
    def mark_key_failed(self, key, temporary=True):
        """علامت‌گذاری کلید به عنوان ناموفق"""
        with self.lock:
            if temporary:
                self.key_usage_count[key] += 1000
            else:
                self.failed_keys.add(key)
    
    def get_stats(self):
        """آمار استفاده از کلیدها"""
        with self.lock:
            active_keys = len(self.api_keys) - len(self.failed_keys)
            total_usage = sum(self.key_usage_count.values())
            return {
                "total_keys": len(self.api_keys),
                "active_keys": active_keys,
                "failed_keys": len(self.failed_keys),
                "total_usage": total_usage
            }

def _is_rate_limit_error(exc: Exception) -> bool:
    msg = str(exc)
    return any(p.search(msg) for p in _rate_limit_patterns)

def extract_retry_delay(error_msg: str) -> int:
    """استخراج زمان انتظار از پیام خطا"""
    retry_match = re.search(r"retry.*?(\d+(?:\.\d+)?)\s*s", error_msg, re.IGNORECASE)
    if retry_match:
        return int(float(retry_match.group(1))) + 5
    return 60

def extract_text_from_pdf(pdf_bytes):
    """استخراج متن از PDF"""
    try:
        doc = fitz.open(stream=pdf_bytes)
        return "".join([page.get_text() for page in doc])
    except Exception as e:
        return ""

def estimate_birth_year_from_text(text):
    """تخمین سال تولد از روی سن"""
    match = re.search(r"(?:سن\s*[:\-]?)?\s*(\d{2})\s*سال", text)
    if match:
        age = int(match.group(1))
        estimated = 1404 - age
        if 1300 <= estimated <= 1404:
            return estimated
    return ""

def clean_year_of_birth(value):
    """پاک‌سازی year_of_birth"""
    try:
        year = float(value)
        year_int = int(round(year))
        return year_int if 1300 <= year_int <= 1404 else ""
    except:
        return ""

def format_job_experience(job_list):
    """تبدیل لیست سوابق شغلی به متن نمایشی"""
    if isinstance(job_list, list):
        return "; ".join([f"{item.get('شرکت', '')} ({item.get('مدت', '')})" for item in job_list])
    return job_list

def format_courses(course_list):
    """تبدیل لیست دوره‌های آموزشی به متن نمایشی"""
    if isinstance(course_list, list):
        return "; ".join([
            f"{c.get('نام دوره', '')}"
            + (f" - {c['مؤسسه']}" if c.get("مؤسسه") else "")
            + (f" ({c['مدت']})" if c.get("مدت") else "")
            for c in course_list
        ])
    return course_list

def process_single_file(file_info, api_manager, max_retries=3):
    """پردازش یک فایل با استفاده از API Manager"""
    filename, pdf_bytes, extracted_text = file_info
    
    prompt = f"{extracted_text}\nاین متن همان PDF است. اطلاعات این متن اولویت دارد. لطفاً اطلاعات خواسته‌شده را مطابق schema زیر استخراج کن.\n\nسوابق شغلی را به صورت لیستی از آبجکت‌ها بده که هر مورد شامل نام شرکت و مدت زمان اشتغال باشد.\nاگر در رزومه به حقوق یا دستمزد اشاره شده بود، بازه حقوق ماهیانه را به صورت عدد ریالی (تومان × 10000) استخراج کن. اگر فقط یک عدد وجود داشت، هر دو مقدار (حداقل و حداکثر) برابر همان عدد باشد."
    
    for attempt in range(max_retries):
        try:
            client, current_key = api_manager.get_available_client()
            
            response = client.models.generate_content(
                model="gemini-2.0-flash",
                contents=[
                    types.Part.from_bytes(data=pdf_bytes, mime_type='application/pdf'),
                    types.Part(text=prompt)
                ],
                config={
                    'response_mime_type': 'application/json',
                    'system_instruction': 'extract asked information from Persian resume',
                    'response_schema': { 
                      "type": "object",
                        "properties": {
                            "نام": {"type": "string", "nullable": False,"description": "extract just first name in persian language."},
                            "نام خانوادگی": {"type": "string", "nullable": False,"description": "extract just family name in persian language"},
                            "شماره تماس": {"type": "string", "nullable": False,"description": "extract just one phone number that begin with 09"},
                            "جنسیت": {"type": "string", "nullable": False,"description": "افراد با جنسیت مذکر را 'آقا'بنویس و جنسیت مونث را 'خانم'بنویس. از نوشتن مرد، مذکر، زن، مونث خودداری کن"},
                            "ایمیل": {"type": "string", "nullable": False,"description":'extract email, prefer gmail if multiple emails exist'},
                            "کانال دریافت رزومه": {"type": "string", "nullable": False,"description": "print 'جاب ویژن' for everyone"},
                            "معرف": {"type": "string", "nullable": False,"description":'این فیلد رو همیشه خالی بذار'},
                            "کارشناسی": {
                                "type": "string", "nullable": True,
                                "description": "لطفاً اطلاعات را به صورت کامل و ساختاریافته بنویس. مثال: 'کارشناسی - مهندسی صنایع - دانشگاه تهران - 1395 تا 1399 - معدل 17.30'"
                            },
                            "کارشناسی ارشد": {
                                "type": "string", "nullable": True,
                                "description": "مثال: 'کارشناسی ارشد - اقتصاد - دانشگاه شهید بهشتی تهران - 1402 تا کنون - معدل 18.02'"
                            },
                            "دکتری": {
                                "type": "string", "nullable": True,
                                "description": "مثال: 'دکتری - مدیریت منابع انسانی - دانشگاه علامه طباطبایی - 1398 تا 1402 - معدل 17.75'"
                            },
                            "رشته تحصیلی": {"type": "string", "nullable": False,"description":'آخرین رشته تحصیلی ای که خونده'},
                            "گرایش تحصیلی": {"type": "string", "nullable": True},
                            "دانشگاه محل تحصیلی": {"type": "string", "nullable": False,"description":'آخرین دانشگاهی که تحصیل کرده'},
                            "نوع دانشگاه آخرین مدرک تحصیلی": {"type": "string", "enum": ["دولتی", "آزاد", "غیر انتفاعی", "پیام نور", "فنی حرفه ای"]},
                            "وضعیت تحصیلی": {"type": "string", "enum": ["فارغ التحصیل کارشناسی ارشد", "فارغ التحصیل دکتری", "دانشجوی کارشناسی", "دانشجوی کارشناسی ارشد", "دانشجوی دکتری"]},
                            "مقطع تحصیلی": {"type": "string", "enum": ["کارشناسی", "دکتری", "کارشناسی ارشد", "کاردانی کارشناسی"]},
                            "نرم افزارها": {"type": "string", "nullable": True},
                            "دوره های آموزشی": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "نام دوره": {"type": "string"},
                                        "مؤسسه": {"type": "string", "nullable": True},
                                        "مدت": {"type": "string", "nullable": True}
                                    },
                                    "required": ["نام دوره"]
                                }
                            },
                            "زبان های خارجی": {"type": "array", "items": {"type": "string"}},
                            "وضعیت خدمت سربازی": {"type": "string", "enum": ["پایان خدمت", "مشمول", "معافیت تحصیلی", "معافیت", "خانم"]},
                            "وضعیت تاهل": {"type": "string", "enum": ["متاهل", "مجرد"]},
                            "year_of_birth": {"type": "number", "nullable": True},
                            "سن": {"type": "number", "nullable": True},
                            "محل سکونت": {"type": "string", "nullable": True},
                            "سوابق شغلی": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "شرکت": {"type": "string"},
                                        "مدت": {"type": "string"}
                                    },
                                    "required": ["شرکت", "مدت"]
                                },
                                "nullable": True
                            },
                            "حداقل حقوق ماهیانه": {"type": "number", "nullable": True},
                            "حداکثر حقوق ماهیانه": {"type": "number", "nullable": True},
                            "فعالیت های داوطلبانه": {"type": "string", "nullable": True},
                            "درباره ی من": {"type": "string", "nullable": True}
                        },
                        "required": [
                            "نام", "نام خانوادگی", "شماره تماس", "جنسیت", "ایمیل", "کانال دریافت رزومه",
                            "رشته تحصیلی", "دانشگاه محل تحصیلی", "نوع دانشگاه آخرین مدرک تحصیلی",
                            "وضعیت تحصیلی", "year_of_birth", "سن", "نرم افزارها",
                            "دوره های آموزشی", "وضعیت خدمت سربازی", "وضعیت تاهل", "مقطع تحصیلی"
                        ]  
                    }
                }
            )
            
            result = json.loads(response.text)
            return {"success": True, "data": result, "filename": filename, "key_used": current_key}
            
        except Exception as e:
            error_msg = str(e)
            
            if _is_rate_limit_error(e):
                api_manager.mark_key_failed(current_key, temporary=True)
                
                if attempt < max_retries - 1:
                    retry_delay = extract_retry_delay(error_msg)
                    time.sleep(min(retry_delay, 10))
                    continue
                else:
                    return {"success": False, "error": f"Rate limit exceeded after {max_retries} attempts", "filename": filename}
            else:
                api_manager.mark_key_failed(current_key, temporary=False)
                
                if attempt < max_retries - 1:
                    time.sleep(2)
                    continue
                else:
                    return {"success": False, "error": str(e), "filename": filename}
    
    return {"success": False, "error": "Max retries exceeded", "filename": filename}

def process_resume_data(row, text):
    """پردازش و تنظیم داده‌های رزومه"""
    
    phone = row.get("شماره تماس", "")
    if phone.startswith("0"):
        row["شماره تماس"] = phone[1:]

    processed_phone = row.get("شماره تماس", "")
    row["شناسه"] = processed_phone if processed_phone else f"ID_{hash(str(row))}"

    row["year_of_birth"] = clean_year_of_birth(row.get("year_of_birth", ""))
    if not row["year_of_birth"]:
        row["year_of_birth"] = estimate_birth_year_from_text(text)

    row["سوابق شغلی"] = format_job_experience(row.get("سوابق شغلی", ""))
    row["دوره های آموزشی"] = format_courses(row.get("دوره های آموزشی", ""))

    reasons = []

    gender = str(row.get("جنسیت", "")).strip()
    degree = str(row.get("مقطع تحصیلی", "")).strip()
    military_status = str(row.get("وضعیت خدمت سربازی", "")).strip()
    max_salary = row.get("حداکثر حقوق ماهیانه", "")

    if "خانم" in gender:
        reasons.append("جنسیت خانم باعث رد شده است.")

    try:
        if max_salary and float(max_salary) > 60_000_000:
            reasons.append("درخواست حقوق بیش از 60 میلیون تومان باعث رد شده است.")
    except:
        pass

    if degree not in ["کارشناسی", "کارشناسی ارشد", "دکتری"]:
        reasons.append("مدرک تحصیلی کمتر از کارشناسی باعث رد شده است.")

    if "مشمول" in military_status:
        reasons.append("مشمول بودن وضعیت سربازی باعث رد شده است.")

    if reasons:
        row["تایید و رد اولیه"] = "رد"
        row["علت رد"] = "؛ ".join(reasons)
        return row, "rejected"
    else:
        row["تایید و رد اولیه"] = "تایید"
        row["علت رد"] = ""
        return row, "approved"

def create_excel_file(all_data):
    """ایجاد فایل Excel با استایل مناسب"""
    df = pd.DataFrame(all_data)
    for col in df.columns:
        df[col] = df[col].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    df = df[[col for col in ORDERED_FIELDS if col in df.columns]]

    output = BytesIO()
    
    base_fields = ["شناسه", "نام", "نام خانوادگی", "شماره تماس", "جنسیت", "ایمیل", "کانال دریافت رزومه", "معرف"]
    base_indexes = [df.columns.get_loc(f) for f in base_fields if f in df.columns]
    if base_indexes:
        base_start = min(base_indexes) + 1
        base_end = max(base_indexes) + 1
    else:
        base_start = 1
        base_end = 8

    check_start = base_end + 1
    check_end = df.shape[1]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="رزومه‌ها", startrow=1)
        workbook = writer.book
        worksheet = writer.sheets["رزومه‌ها"]

        worksheet.insert_rows(1)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=base_end)
        worksheet.merge_cells(start_row=1, start_column=base_end + 1, end_row=1, end_column=check_end)
        worksheet.cell(row=1, column=1).value = "مشخصات پایه و معرف"
        worksheet.cell(row=1, column=base_end + 1).value = "بررسی رزومه"
            
        base_fill = PatternFill(start_color="C2E0FF", end_color="C2E0FF", fill_type="solid")
        check_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        group_font = Font(bold=True, size=13)
        group_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, base_end + 1):
            worksheet.cell(row=1, column=col).fill = base_fill
            worksheet.cell(row=1, column=col).font = group_font
            worksheet.cell(row=1, column=col).alignment = group_alignment
        for col in range(base_end + 1, check_end + 1):
            worksheet.cell(row=1, column=col).fill = check_fill
            worksheet.cell(row=1, column=col).font = group_font
            worksheet.cell(row=1, column=col).alignment = group_alignment

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        for cell in worksheet[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        approve_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        reject_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        try:
            status_col_idx = ORDERED_FIELDS.index("تایید و رد اولیه")
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                if len(row) > status_col_idx:
                    status = row[status_col_idx].value
                    if status == "تایید":
                        for cell in row:
                            cell.fill = approve_fill
                    elif status == "رد":
                        for cell in row:
                            cell.fill = reject_fill
        except (ValueError, IndexError):
            pass

        for idx, col in enumerate(worksheet.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output.seek(0)
    return output

def main():
    # هدر مدرن
    st.markdown('''
    <div class="modern-header">
        <h1>✨ پردازشگر هوشمند رزومه</h1>
        <p>پردازش سریع و دقیق رزومه‌ها با هوش مصنوعی</p>
    </div>
    ''', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ تنظیمات")
        
        st.markdown("---")
        st.markdown("#### 🔑 کلیدهای API")
        
        key_input_method = st.radio(
            "روش ورود کلیدها:",
            ["کلیدهای پیش‌فرض", "کلیدهای سفارشی"],
            help="می‌توانید از کلیدهای پیش‌فرض استفاده کنید یا کلیدهای خود را وارد کنید"
        )
        
        if key_input_method == "کلیدهای پیش‌فرض":
            api_keys = DEFAULT_GENAI_KEYS
            st.success(f"✅ {len(api_keys)} کلید لود شد")
        else:
            custom_keys_text = st.text_area(
                "کلیدهای API (هر کدام در یک خط):",
                height=120,
                placeholder="AIzaSy...\nAIzaSy..."
            )
            
            if custom_keys_text:
                api_keys = [key.strip() for key in custom_keys_text.split('\n') if key.strip()]
                st.success(f"✅ {len(api_keys)} کلید سفارشی")
            else:
                api_keys = DEFAULT_GENAI_KEYS
                st.warning("⚠️ از کلیدهای پیش‌فرض استفاده می‌شود")

        st.markdown("---")
        st.markdown("#### 🌐 پروکسی")
        
        use_proxy = st.checkbox("فعال‌سازی پروکسی")
        
        if use_proxy:
            proxy_url = st.text_input(
                "آدرس پروکسی:",
                value="",
                placeholder="http://your-proxy-address:port"
            )
            if proxy_url:
                os.environ['HTTP_PROXY'] = proxy_url
                os.environ['HTTPS_PROXY'] = proxy_url
                st.success("✅ پروکسی فعال شد")
        else:
            os.environ.pop('HTTP_PROXY', None)
            os.environ.pop('HTTPS_PROXY', None)

        st.markdown("---")
        st.markdown("#### ⚡ پردازش")
        
        max_workers = st.slider(
            "Thread های همزمان:",
            1, min(len(api_keys), 10), 
            min(len(api_keys), 5),
            help="تعداد فایل‌هایی که همزمان پردازش می‌شوند"
        )
        
        max_retries = st.slider(
            "تلاش مجدد:",
            1, 5, 3,
            help="تعداد دفعات تلاش مجدد در صورت خطا"
        )

    # تب‌های اصلی
    tab1, tab2, tab3, tab4 = st.tabs([
        "📤 آپلود و پردازش",
        "📊 نتایج", 
        "📈 آمار API",
        "📚 راهنما"
    ])
    
    with tab1:
        st.markdown('<div class="modern-card">', unsafe_allow_html=True)
        st.markdown("### 📤 آپلود فایل‌های رزومه")
        
        uploaded_files = st.file_uploader(
            "فایل‌های PDF را بکشید و رها کنید یا کلیک کنید",
            type=['pdf'],
            accept_multiple_files=True,
            help="حداکثر تا 100 فایل به صورت همزمان"
        )
        
        if uploaded_files:
            st.markdown(f'<div class="success-box-modern">✅ {len(uploaded_files)} فایل آپلود شد</div>', unsafe_allow_html=True)
            
            with st.expander("📋 فایل‌های آپلود شده", expanded=False):
                for i, file in enumerate(uploaded_files, 1):
                    st.write(f"**{i}.** {file.name} ({file.size:,} بایت)")
            
            st.markdown(f'<div class="info-box-modern">🚀 پردازش با {max_workers} Thread و {len(api_keys)} کلید API</div>', unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("🚀 شروع پردازش", type="primary", use_container_width=True):
                    process_files_parallel(uploaded_files, api_keys, max_workers, max_retries)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tab2:
        st.markdown("### 📊 نتایج پردازش")
        
        if "processing_results" in st.session_state and st.session_state.processing_results:
            display_results()
        else:
            st.markdown('<div class="info-box-modern">🔍 هنوز فایلی پردازش نشده است</div>', unsafe_allow_html=True)
    
    with tab3:
        st.markdown("### 📈 آمار استفاده از API")
        
        if "api_stats" in st.session_state and st.session_state.api_stats:
            display_api_stats()
        else:
            st.markdown('<div class="info-box-modern">🔍 آماری موجود نیست</div>', unsafe_allow_html=True)
    
    with tab4:
        display_help()

def process_files_parallel(uploaded_files, api_keys, max_workers, max_retries):
    """پردازش فایل‌های آپلود شده"""
    
    st.markdown('<div class="info-box-modern">🔄 در حال شروع پردازش...</div>', unsafe_allow_html=True)
    
    api_manager = APIKeyManager(api_keys)
    
    file_data = []
    for uploaded_file in uploaded_files:
        pdf_bytes = uploaded_file.read()
        extracted_text = extract_text_from_pdf(pdf_bytes)
        
        if extracted_text.strip():
            file_data.append((uploaded_file.name, pdf_bytes, extracted_text))
        else:
            st.markdown(f'<div class="warning-box-modern">⚠️ فایل {uploaded_file.name} قابل خواندن نیست</div>', unsafe_allow_html=True)
    
    if not file_data:
        st.markdown('<div class="error-box-modern">❌ هیچ فایل قابل پردازشی وجود ندارد</div>', unsafe_allow_html=True)
        return
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    processing_stats = {
        "total": len(file_data),
        "processed": 0,
        "failed": 0,
        "approved": 0,
        "rejected": 0,
        "start_time": time.time()
    }
    
    metrics_container = st.container()
    details_container = st.container()
    
    with metrics_container:
        col1, col2, col3, col4, col5 = st.columns(5)
        metric_total = col1.empty()
        metric_processed = col2.empty()
        metric_approved = col3.empty()
        metric_rejected = col4.empty()
        metric_speed = col5.empty()
    
    with details_container:
        details_expander = st.expander("🔍 جزئیات لحظه‌ای", expanded=True)
        details_text = details_expander.empty()
    
    processing_details = []
    all_data = []
    failed_files = []
    
    def update_ui():
        elapsed_time = time.time() - processing_stats["start_time"]
        speed = processing_stats["processed"] / max(elapsed_time, 1) * 60
        
        metric_total.markdown(f'''
        <div class="metric-modern">
            <p>کل فایل‌ها</p>
            <h3>{processing_stats["total"]}</h3>
        </div>
        ''', unsafe_allow_html=True)
        
        metric_processed.markdown(f'''
        <div class="metric-modern">
            <p>پردازش شده</p>
            <h3>{processing_stats["processed"]}</h3>
        </div>
        ''', unsafe_allow_html=True)
        
        metric_approved.markdown(f'''
        <div class="metric-modern">
            <p>تایید شده</p>
            <h3>{processing_stats["approved"]}</h3>
        </div>
        ''', unsafe_allow_html=True)
        
        metric_rejected.markdown(f'''
        <div class="metric-modern">
            <p>رد شده</p>
            <h3>{processing_stats["rejected"]}</h3>
        </div>
        ''', unsafe_allow_html=True)
        
        metric_speed.markdown(f'''
        <div class="metric-modern">
            <p>سرعت (فایل/دقیقه)</p>
            <h3>{speed:.1f}</h3>
        </div>
        ''', unsafe_allow_html=True)
        
        details_text.text("\n".join(processing_details[-15:]))
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_file = {
            executor.submit(process_single_file, file_info, api_manager, max_retries): file_info[0] 
            for file_info in file_data
        }
        
        for future in as_completed(future_to_file):
            filename = future_to_file[future]
            
            try:
                result = future.result()
                
                if result["success"]:
                    model_output = result["data"]
                    row = {field: model_output.get(field, "") for field in ORDERED_FIELDS}
                    
                    if "شناسه" not in row or not row["شناسه"]:
                        row["شناسه"] = model_output.get("شماره تماس", f"ID_{len(all_data)+1}")
                    
                    file_text = ""
                    for file_info in file_data:
                        if file_info[0] == filename:
                            file_text = file_info[2]
                            break
                    
                    processed_row, status = process_resume_data(row, file_text)
                    all_data.append(processed_row)
                    
                    processing_stats["processed"] += 1
                    
                    if status == "approved":
                        processing_stats["approved"] += 1
                        processing_details.append(f"✅ تایید: {processed_row.get('نام', '')} {processed_row.get('نام خانوادگی', '')}")
                    else:
                        processing_stats["rejected"] += 1
                        processing_details.append(f"❌ رد: {processed_row.get('نام', '')} {processed_row.get('نام خانوادگی', '')}")
                else:
                    failed_files.append(filename)
                    processing_stats["failed"] += 1
                    processing_details.append(f"❌ خطا: {filename}")
                
            except Exception as e:
                failed_files.append(filename)
                processing_stats["failed"] += 1
                processing_details.append(f"❌ خطای غیرمنتظره: {filename}")
            
            progress = (processing_stats["processed"] + processing_stats["failed"]) / processing_stats["total"]
            progress_bar.progress(progress)
            status_text.text(f"🔄 پردازش: {processing_stats['processed'] + processing_stats['failed']}/{processing_stats['total']}")
            
            update_ui()
    
    progress_bar.progress(1.0)
    status_text.text("✅ پردازش کامل شد!")
    
    total_time = time.time() - processing_stats["start_time"]
    
    st.session_state.processing_results = {
        "data": all_data,
        "stats": processing_stats,
        "failed_files": failed_files,
        "processing_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_time": total_time
    }
    
    st.session_state.api_stats = api_manager.get_stats()
    
    st.markdown(f'<div class="success-box-modern">🎉 پردازش در {total_time:.1f} ثانیه تکمیل شد!</div>', unsafe_allow_html=True)

def display_results():
    """نمایش نتایج پردازش"""
    
    results = st.session_state.processing_results
    
    st.markdown("#### آمار کلی")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f'''
        <div class="metric-modern">
            <p>کل فایل‌ها</p>
            <h3>{results['stats']['total']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col2:
        st.markdown(f'''
        <div class="metric-modern">
            <p>پردازش موفق</p>
            <h3>{results['stats']['processed']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col3:
        st.markdown(f'''
        <div class="metric-modern">
            <p>تایید شده</p>
            <h3>{results['stats']['approved']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col4:
        st.markdown(f'''
        <div class="metric-modern">
            <p>رد شده</p>
            <h3>{results['stats']['rejected']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    if results['data']:
        st.markdown("#### توزیع نتایج")
        
        chart_data = pd.DataFrame({
            'وضعیت': ['تایید شده', 'رد شده'],
            'تعداد': [results['stats']['approved'], results['stats']['rejected']]
        })
        
        st.bar_chart(chart_data.set_index('وضعیت'))
    
    if 'total_time' in results:
        st.markdown("#### عملکرد پردازش")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("⏱️ زمان کل", f"{results['total_time']:.1f} ثانیه")
        
        with col2:
            avg_time = results['total_time'] / results['stats']['total'] if results['stats']['total'] > 0 else 0
            st.metric("📊 متوسط", f"{avg_time:.1f} ث/فایل")
        
        with col3:
            speed = results['stats']['total'] / results['total_time'] * 60 if results['total_time'] > 0 else 0
            st.metric("🚀 سرعت", f"{speed:.1f} فایل/دقیقه")
    
    if results['data']:
        st.markdown("#### داده‌های استخراج شده")
        
        df = pd.DataFrame(results['data'])
        
        if 'شناسه' not in df.columns:
            df['شناسه'] = df['شماره تماس'] if 'شماره تماس' in df.columns else [f"ID_{i+1}" for i in range(len(df))]
        
        filter_status = st.selectbox(
            "فیلتر بر اساس وضعیت:",
            ["همه", "تایید شده", "رد شده"]
        )
        
        if filter_status == "تایید شده":
            df_filtered = df[df["تایید و رد اولیه"] == "تایید"]
        elif filter_status == "رد شده":
            df_filtered = df[df["تایید و رد اولیه"] == "رد"]
        else:
            df_filtered = df
        
        st.dataframe(df_filtered, use_container_width=True, height=400)
        
        if st.button("📥 دانلود فایل Excel"):
            excel_data = results['data'].copy() if isinstance(results['data'], list) else results['data']
            if isinstance(excel_data, list):
                for item in excel_data:
                    if 'شناسه' not in item or not item['شناسه']:
                        item['شناسه'] = item.get('شماره تماس', f"ID_{excel_data.index(item)+1}")
            
            excel_file = create_excel_file(excel_data)
            
            st.download_button(
                label="💾 دانلود Excel",
                data=excel_file,
                file_name=f"resume_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    if results['failed_files']:
        st.markdown("#### ⚠️ فایل‌های پردازش نشده")
        st.markdown('<div class="error-box-modern">', unsafe_allow_html=True)
        for failed_file in results['failed_files']:
            st.write(f"• {failed_file}")
        st.markdown('</div>', unsafe_allow_html=True)

def display_api_stats():
    """نمایش آمار استفاده از API"""
    
    stats = st.session_state.api_stats
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f'''
        <div class="metric-modern">
            <p>کل کلیدها</p>
            <h3>{stats['total_keys']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col2:
        st.markdown(f'''
        <div class="metric-modern">
            <p>فعال</p>
            <h3>{stats['active_keys']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col3:
        st.markdown(f'''
        <div class="metric-modern">
            <p>غیرفعال</p>
            <h3>{stats['failed_keys']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    with col4:
        st.markdown(f'''
        <div class="metric-modern">
            <p>کل درخواست‌ها</p>
            <h3>{stats['total_usage']}</h3>
        </div>
        ''', unsafe_allow_html=True)
    
    if stats['total_keys'] > 0:
        chart_data = pd.DataFrame({
            'وضعیت': ['فعال', 'غیرفعال'],
            'تعداد': [stats['active_keys'], stats['failed_keys']]
        })
        
        st.bar_chart(chart_data.set_index('وضعیت'))

def display_help():
    """نمایش راهنما"""
    
    help_sections = {
        "🚀 پردازش موازی": """
        **مزایای پردازش موازی:**
        - استفاده همزمان از چندین کلید API
        - سرعت پردازش بالاتر (تا 5-10 برابر)
        - مدیریت خودکار کلیدهای ناموفق
        - بهره‌وری بهتر از منابع
        """,
        
        "🔑 مدیریت هوشمند کلیدها": """
        **ویژگی‌های مدیر کلیدها:**
        - توزیع یکنواخت بار بین کلیدها
        - تشخیص خودکار کلیدهای ناموفق
        - مدیریت محدودیت‌های موقت و دائم
        """,
        
        "⚙️ تنظیمات بهینه": """
        **نکات مهم:**
        - Thread کمتر از تعداد کلیدها
        - برای اتصال سریع: 3-5 Thread
        - برای اتصال آهسته: 1-2 Thread
        - کلیدهای بیشتر = سرعت بالاتر
        """
    }
    
    for section_title, section_content in help_sections.items():
        with st.expander(section_title, expanded=False):
            st.markdown(section_content)

if __name__ == "__main__":
    if "processing_results" not in st.session_state:
        st.session_state.processing_results = None
    
    if "api_stats" not in st.session_state:
        st.session_state.api_stats = None
    
    main()
